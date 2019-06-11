#!/usr/bin/env python
import csv
import sys
import os
from openpyxl import load_workbook, Workbook
from shutil import copy2

samples = {}
assaylist = []

def main():
    # path to folder for platemap
    path = getpath()
    # gets platemap from folder
    path1 = getplatemap(path)
    # gets specific sheet from platemap
    platemap1 = platemapsheet1(path1)
    # gets all assays and initial samplecount
    getassays(platemap1)
    # checks for empty assays and adds overage
    samplecount = checksamples()
    # location of mastermix recipee template
    mmtemplate = 'C:\\Users\\Derrik\\Desktop\\coding\\cs50\\mastermixer\\template.xlsx'
    # gets PCR reagents
    reagents = open_reagent_list()
    # outputs assays, samplecount, and reagents to template copy
    MMoutput(mmtemplate, path, reagents)
    

def getpath():
    # path to folder for platemap
    while True:
        path = input("Enter the path of your platemap file: ")
        if not os.path.isdir(path):
            print("Not a valid directory")
        else:
            break
    return path


def getplatemap(path):
    # gets platemap from folder
    filelist= []
    for file in os.listdir(path):
        filelist.append(file)
    print('Here is a list of the files in that directory.')
    print(filelist)
    platemap = fileselect(filelist)
    path1 =  path + "\\" + platemap
    return path1


def fileselect(filelist):
    # gets user input for which file in the folder is the platemap
    while True: 
        try:
            filenum = int(input('Which would you like to use? (enter a number corresponding to the order of the files):'))
            platemap = filelist[filenum - 1]
            print("You selected '{}'. Are you sure?".format(platemap))
            yesno = input(':')
            yeslist = ('Y', 'y', 'yes', 'Yes', 'n', 'N', 'No', 'no')
            if yesno in yeslist[:3]:
                return platemap
            if yesno not in yeslist:
                print("this is a yes or no question")
        except TypeError:
            print("Use a number dummy")
        except ValueError:
            print("Use a number dummy")
        except IndexError:
            print("Count much? Use a number that refers to one of the files")        


def platemapsheet1(path1):
    # gets specific sheet from platemap
    wb = load_workbook(filename=path1, data_only=True)
    ws = wb.worksheets[0]
    return ws

           
def getassays(ws):
    # gets all assays and initial samplecount
    # Hardcoded the rows I want to examine. I couldn't figure out a consistent rule to avoid hardcoding these
    assay_rows = [2, 15, 18, 31, 34, 47, 50, 63, 66, 79, 82, 95, 98, 111, 114, 127,
                   130, 143, 146, 159, 162, 175, 178, 191, 194, 207, 210, 223, 226, 239, 242, 255]
    bottom_rows = [15, 31, 47, 63, 79, 95, 111, 127, 143, 159, 175, 191, 207, 223, 239, 255]
    rowcounter = 0
    assaycounter = 0
    columntitles = []
    bottomrowcounter = 0
    botrowlen = len(bottom_rows)
    for row in ws.values:
        # Resets a column value for each new row. The column value allows me to iterate trough each member of the list I made from the row
        # I start at 4 because I know the values I care about will start in column 4
        columncounter = 3
        rowcounter += 1
        if rowcounter in assay_rows:
            columntitles.clear()     
            while True:
                cell = '{}{}'.format(chr(64 + columncounter), rowcounter)
                # Will add the value in the current cell to the assay list if it is not already present in the list
                if ws[cell].value not in assaylist and ws[cell].value != None:
                    assaylist.append(ws[cell].value)
                    if ws[cell].value not in samples:
                        samples[(ws[cell].value)] = []
                columntitles.append(ws[cell].value)    
                columncounter += 1
                if columncounter == 15:
                    break
        if (rowcounter - 7) % 16 == 0:
            for a in range(12):
                # each column
                temprowcounter = rowcounter
                for b in range(8):
                    # each row
                    cell = '{}{}'.format(chr(64 + columncounter), temprowcounter)
                    if columntitles[columncounter - 3] != None:
                        if ws[cell].value == ' ' or ws[cell].value == None:
                            placeholder = 0
                        else:
                            samples[(columntitles[columncounter - 3])].append(ws[cell].value)
                        if ws[cell].value == 'RNTC_NTC_A_1_1':
                            # checks if there is another assay tucked under the end of this one
                            cell1 = ('{}{}'.format(chr(64 + columncounter), (bottom_rows[bottomrowcounter])))
                            str1 = ws[cell1].value
                            str2 = columntitles[columncounter - 3]
                            if str1 != str2 and str1 != None:
                                columntitles[columncounter - 3] = str1
                                samples[str1] = [] 
                    temprowcounter += 1
                columncounter += 1
        if rowcounter > bottom_rows[bottomrowcounter]:
                bottomrowcounter += 1
        if rowcounter > bottom_rows[botrowlen - 1]:
            break
        
            
def checksamples():
    # checks for empty assays and adds overage
    badassays = []
    samplecount = {}
    for assays in samples:
        if len(samples[assays]) == 0:
            badassays.append(assays)
    print("The following strings were in places where assays should be, but they're probably not assays.")
    print(badassays)
    print('Should they be removed?')
    answers = ['Y', 'y', 'yes', 'yes','n','N','no','No']
    response = 'a'
    while response not in answers:
        response = input(':')
    if response in answers[:3]:
        for assays in badassays:
            assaylist.remove(assays)
    for assays in assaylist:
        samplecount[assays] = 0
        n = len(samples[assays])
        if n / 10 < 4:
            samplecount[assays] = round(n * 1.1)    
        else:
            samplecount[assays] = n + 4
        if n < 5:
            samplecount[assays] = n + 0.5

    
def open_reagent_list():
    # gets PCR reagents
    reagents = {}
    with open('assaydictionary.csv', 'r') as a:
        reagent_list = csv.reader(a)
        for rows in reagent_list:
            if rows[0] in assaylist:
                reagents[rows[0]] = rows[1]
    return reagents


def edittemplate(ws, reagents):
    # all of the cell editing
    rowcounter = 1
    current_assay = 0
    for row in ws.iter_rows():
        columncounter = 1
        if rowcounter % 7 == 0:
            for x in range(3):
                ws.cell(row=rowcounter, column=columncounter).value = assaylist[current_assay]
                ws.cell(row=(rowcounter + 2), column=columncounter).value = reagents.get(assaylist[current_assay], "Not Found")
                if assaylist[current_assay] in samples:
                    #sample count
                    ws.cell(row=rowcounter, column=(columncounter + 3)).value = samplecount[assaylist[current_assay]]
                    #primer volume
                    ws.cell(row=(rowcounter + 5), column=(columncounter + 1)).value = 2
                else:
                    print("Error: {} slipped through the cracks. Is there something weird about this assay on the platemap?".format(assaylist[current_assay]))
                if reagents.get(assaylist[current_assay]) == 'Zymo PCR Multiplex':
                    ws.cell(row=(rowcounter + 3), column=columncounter).value = 'DMSO'
                    ws.cell(row=(rowcounter + 3), column=(columncounter + 1)).value = .5
                else:
                    ws.cell(row=(rowcounter + 3), column=columncounter).value = None
                    ws.cell(row=(rowcounter + 3), column=(columncounter + 2)).value = None
                if assaylist[current_assay] == 'ATP7B_112GA_RD_2':
                    ws.cell(row=(rowcounter + 5), column=(columncounter + 1)).value = 3.5
                current_assay += 1
                columncounter += 4
                if current_assay == len(assaylist):
                    return


def MMoutput(template, path, reagents):
    # combine all info on template and save to folder with platemap
    wb = load_workbook(template)
    ws = wb.active
    edittemplate(ws, reagents)   
    wb.save('Secondary_PCR_Mastermixes.xlsx')
    copy2(template, path)


if __name__ == "__main__":
    main()